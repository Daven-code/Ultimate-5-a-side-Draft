"""
Create a separate JSON file containing ONLY the WorldCup2026 sheet players.

Typical use from your project folder:
  python convert_worldcup_excel_to_players_json.py Ultimate5Aside_Player_Pool.xlsm

Optional:
  python convert_worldcup_excel_to_players_json.py Ultimate5Aside_Player_Pool.xlsm players_worldcup2026.json

This does NOT use the Data tab, so it will not affect the normal game player pool.
"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

workbook = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('Ultimate5Aside_Player_Pool.xlsm')
output = Path(sys.argv[2]) if len(sys.argv) > 2 else Path('players_worldcup2026.json')
sheet_name = 'WorldCup2026'

if not workbook.exists():
    raise SystemExit(f'File not found: {workbook}')

wb = load_workbook(workbook, data_only=True)
if sheet_name not in wb.sheetnames:
    raise SystemExit(f'Workbook must contain a sheet named {sheet_name}')

ws = wb[sheet_name]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
required = ['Include', 'Player', 'Game_Year', 'Rating_OVR', 'Position', 'Main_Position', 'Club', 'Nation', 'Squad_Nation', 'Challenge_Mode']
missing = [h for h in required if h not in headers]
if missing:
    raise SystemExit(f'Missing expected columns: {missing}')

idx = {h: headers.index(h) + 1 for h in required}
players = []
blank_ratings = []

for r in range(2, ws.max_row + 1):
    include = ws.cell(r, idx['Include']).value
    player = ws.cell(r, idx['Player']).value
    if not player or str(include).strip().upper() not in ('Y', 'YES', 'TRUE', '1'):
        continue

    rating = ws.cell(r, idx['Rating_OVR']).value
    if rating in (None, ''):
        blank_ratings.append(f"row {r}: {player} ({ws.cell(r, idx['Squad_Nation']).value})")
        continue

    players.append({
        'Player': player,
        'Game_Year': int(ws.cell(r, idx['Game_Year']).value),
        'Rating_OVR': int(rating),
        'Position': ws.cell(r, idx['Position']).value,
        'Main_Position': ws.cell(r, idx['Main_Position']).value,
        'Club': ws.cell(r, idx['Club']).value or '',
        'Nation': ws.cell(r, idx['Nation']).value,
        'Squad_Nation': ws.cell(r, idx['Squad_Nation']).value,
        'Challenge_Mode': ws.cell(r, idx['Challenge_Mode']).value,
    })

output.write_text(json.dumps(players, ensure_ascii=False, indent=2), encoding='utf-8')
print(f'Created {output} with {len(players)} World Cup players')
if blank_ratings:
    print(f'Skipped {len(blank_ratings)} rows with blank ratings:')
    for item in blank_ratings[:50]:
        print(' - ' + item)
    if len(blank_ratings) > 50:
        print(f' ... plus {len(blank_ratings)-50} more')
