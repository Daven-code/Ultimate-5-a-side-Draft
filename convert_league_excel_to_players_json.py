"""
Convert the League Legends Excel workbook into league_players.json for Ultimate 5-a-side.

Default input workbook:
  top_200_players_five_leagues.xlsx

Default output:
  league_players.json

Basic usage from your project folder:
  python convert_league_excel_to_players_json.py

Or specify files explicitly:
  python convert_league_excel_to_players_json.py top_200_players_five_leagues.xlsx league_players.json

Useful options:
  python convert_league_excel_to_players_json.py --top-n 200
  python convert_league_excel_to_players_json.py --include-all
  python convert_league_excel_to_players_json.py --sheet "Premier League" --sheet "La Liga"

Notes:
- Reads the main league tabs from top_200_players_five_leagues.xlsx.
- Exports the JSON shape used by the League Legends game mode.
- Uses the workbook rating if Overall Rating is present, otherwise calculates:
  Prime Score 50% + Legacy Score 35% + Longevity Score 15%.
- Defaults to the top 100 players per league to match the current league_players.json size.
- Position multipliers are generated from the listed primary position, using the best multiplier
  across composite positions such as ST/LW, CM/AM, RB/CM, etc.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

DEFAULT_WORKBOOK = "top_200_players_five_leagues.xlsx"
DEFAULT_OUTPUT = "league_players.json"
DEFAULT_TOP_N = 100
DEFAULT_LEAGUE_SHEETS = [
    "Premier League",
    "La Liga",
    "Serie A",
    "Ligue 1",
    "Bundesliga",
]

# Base out-of-position multipliers used by the League Legends mode.
# FWD and ST are exported separately for compatibility with the current game code.
BASE_POSITION_MULTIPLIERS: dict[str, dict[str, float]] = {
    "GK": {"DEF": 0.0, "MID": 0.0, "FWD": 0.0, "ST": 0.0},
    "CB": {"DEF": 1.0, "MID": 0.82, "FWD": 0.55, "ST": 0.55},
    "SW": {"DEF": 1.0, "MID": 0.85, "FWD": 0.58, "ST": 0.58},
    "SWEEPER": {"DEF": 1.0, "MID": 0.85, "FWD": 0.58, "ST": 0.58},
    "LB": {"DEF": 1.0, "MID": 0.82, "FWD": 0.65, "ST": 0.65},
    "RB": {"DEF": 1.0, "MID": 0.82, "FWD": 0.65, "ST": 0.65},
    "FB": {"DEF": 1.0, "MID": 0.82, "FWD": 0.65, "ST": 0.65},
    "LWB": {"DEF": 1.0, "MID": 0.86, "FWD": 0.72, "ST": 0.72},
    "RWB": {"DEF": 1.0, "MID": 0.86, "FWD": 0.72, "ST": 0.72},
    "DM": {"DEF": 0.95, "MID": 1.0, "FWD": 0.70, "ST": 0.70},
    "CDM": {"DEF": 0.95, "MID": 1.0, "FWD": 0.70, "ST": 0.70},
    "CM": {"DEF": 0.85, "MID": 1.0, "FWD": 0.80, "ST": 0.80},
    "LM": {"DEF": 0.75, "MID": 1.0, "FWD": 0.86, "ST": 0.86},
    "RM": {"DEF": 0.75, "MID": 1.0, "FWD": 0.86, "ST": 0.86},
    "LWM": {"DEF": 0.75, "MID": 0.85, "FWD": 0.85, "ST": 0.85},
    "RWM": {"DEF": 0.75, "MID": 0.85, "FWD": 0.85, "ST": 0.85},
    "AM": {"DEF": 0.65, "MID": 1.0, "FWD": 0.90, "ST": 0.90},
    "CAM": {"DEF": 0.65, "MID": 1.0, "FWD": 0.90, "ST": 0.90},
    "LW": {"DEF": 0.65, "MID": 0.82, "FWD": 1.0, "ST": 1.0},
    "RW": {"DEF": 0.65, "MID": 0.82, "FWD": 1.0, "ST": 1.0},
    "W": {"DEF": 0.65, "MID": 0.82, "FWD": 1.0, "ST": 1.0},
    "SS": {"DEF": 0.58, "MID": 0.84, "FWD": 0.98, "ST": 0.98},
    "CF": {"DEF": 0.55, "MID": 0.75, "FWD": 1.0, "ST": 1.0},
    "ST": {"DEF": 0.55, "MID": 0.75, "FWD": 1.0, "ST": 1.0},
    "FW": {"DEF": 0.58, "MID": 0.78, "FWD": 0.98, "ST": 0.98},
}

MAIN_POSITION_TIE_ORDER = ["DEF", "MID", "FWD"]


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def normalise_header(value: Any) -> str:
    return clean_text(value).casefold()


def to_int(value: Any, fallback: int | None = None) -> int:
    if value in (None, ""):
        if fallback is None:
            raise ValueError("Missing integer value")
        return fallback
    try:
        return int(value)
    except (TypeError, ValueError):
        return int(float(value))


def round_half_up(value: float) -> int:
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def rating_from_scores(row: dict[str, Any]) -> int:
    overall = row.get("Overall Rating")
    if overall not in (None, ""):
        try:
            value = float(overall)
            if math.isfinite(value) and value > 0:
                return round_half_up(value)
        except (TypeError, ValueError):
            pass

    prime = float(row.get("Prime Score") or 0)
    legacy = float(row.get("Legacy Score") or 0)
    longevity = float(row.get("Longevity Score") or 0)
    calculated = (prime * 0.50) + (legacy * 0.35) + (longevity * 0.15)
    if calculated <= 0:
        raise ValueError(f"Cannot calculate rating for {row.get('Player')!r}")
    return round_half_up(calculated)


def position_components(position_text: str) -> list[str]:
    text = clean_text(position_text).upper()
    text = text.replace("RIGHT", "R").replace("LEFT", "L")
    parts = re.split(r"[^A-Z]+", text)
    components: list[str] = []
    for part in parts:
        if not part:
            continue
        aliases = {
            "FORWARD": "FW",
            "FORWARDS": "FW",
            "WINGER": "W",
            "WINGERS": "W",
        }
        part = aliases.get(part, part)
        # Generic FB/W can overstate versatile text such as CM/FB/W.
        # Keep them only if they are the only usable component.
        if part in {"FB", "W", "SWEEPER"}:
            continue
        if part in BASE_POSITION_MULTIPLIERS:
            components.append(part)
    return components or ["FW"]


def build_multipliers(position_text: str) -> dict[str, float]:
    components = position_components(position_text)
    if "GK" in components:
        return {"DEF": 0, "MID": 0, "FWD": 0, "ST": 0}

    result = {"DEF": 0.0, "MID": 0.0, "FWD": 0.0, "ST": 0.0}
    for component in components:
        multipliers = BASE_POSITION_MULTIPLIERS.get(component, BASE_POSITION_MULTIPLIERS["FW"])
        for key in result:
            result[key] = max(result[key], float(multipliers[key]))

    # FWD and ST should stay aligned for the current frontend logic.
    result["FWD"] = result["ST"]
    return {key: int(value) if value in (0.0, 1.0) else value for key, value in result.items()}


def main_position_from_multipliers(position_text: str, multipliers: dict[str, float]) -> str:
    components = position_components(position_text)
    if "GK" in components:
        return "GK"
    if components and all(component in {"LWM", "RWM"} for component in components):
        return "FWD"

    best_value = max(float(multipliers[role]) for role in MAIN_POSITION_TIE_ORDER)
    for role in MAIN_POSITION_TIE_ORDER:
        if float(multipliers[role]) == best_value:
            return role
    return "FWD"


def find_column(headers: list[str], possible_names: Iterable[str]) -> str:
    lookup = {normalise_header(header): header for header in headers if header not in (None, "")}
    for name in possible_names:
        found = lookup.get(normalise_header(name))
        if found:
            return found
    raise KeyError(f"Could not find any of these columns: {list(possible_names)}")


def rows_from_sheet(ws) -> tuple[list[dict[str, Any]], dict[str, str]]:
    raw_headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    headers = [clean_text(h) for h in raw_headers]
    header_map = {
        "rank": find_column(headers, ["Rank"]),
        "player": find_column(headers, ["Player"]),
        "position": find_column(headers, ["Primary Position", "Position"]),
        "club": find_column(headers, ["Main League Club(s)", "Main Premier League Club(s)", "Club"]),
        "club_spell": find_column(headers, ["League Club Spell(s)", "Club Spell", "Club_Spell"]),
        "prime": find_column(headers, ["Prime Score"]),
        "legacy": find_column(headers, ["Legacy Score"]),
        "longevity": find_column(headers, ["Longevity Score"]),
        "overall": find_column(headers, ["Overall Rating", "Rating_OVR", "Rating"]),
        "rationale": find_column(headers, ["Short rationale", "Short Rationale", "Short_Rationale"]),
    }

    output: list[dict[str, Any]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if not clean_text(row.get(header_map["player"])):
            continue
        row["__row_number"] = row_number
        output.append(row)
    return output, header_map


def build_player(row: dict[str, Any], header_map: dict[str, str], league: str, source_name: str) -> dict[str, Any]:
    position = clean_text(row.get(header_map["position"]))
    multipliers = build_multipliers(position)
    main_position = main_position_from_multipliers(position, multipliers)

    mapped_row = {
        "Player": clean_text(row.get(header_map["player"])),
        "Prime Score": row.get(header_map["prime"]),
        "Legacy Score": row.get(header_map["legacy"]),
        "Longevity Score": row.get(header_map["longevity"]),
        "Overall Rating": row.get(header_map["overall"]),
    }

    return {
        "Player": mapped_row["Player"],
        "Rank": to_int(row.get(header_map["rank"]), fallback=0),
        "Game_Year": 0,
        "Rating_OVR": rating_from_scores(mapped_row),
        "Position": position,
        "Main_Position": main_position,
        "Club": clean_text(row.get(header_map["club"])),
        "Nation": "",
        "League": league,
        "Club_Spell": clean_text(row.get(header_map["club_spell"])),
        "Short_Rationale": clean_text(row.get(header_map["rationale"])),
        "Position_Multipliers": multipliers,
        "Source": source_name,
        "Notes": "League Legends Challenge player pool",
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert top_200_players_five_leagues.xlsx into league_players.json."
    )
    parser.add_argument("workbook", nargs="?", default=DEFAULT_WORKBOOK, help="Input workbook path")
    parser.add_argument("output", nargs="?", default=DEFAULT_OUTPUT, help="Output JSON path")
    parser.add_argument(
        "--sheet",
        action="append",
        dest="sheets",
        help="League sheet to export. Can be supplied multiple times. Defaults to all five league sheets.",
    )
    parser.add_argument(
        "--top-n",
        type=int,
        default=DEFAULT_TOP_N,
        help=f"Number of players per league sheet to export. Default: {DEFAULT_TOP_N}.",
    )
    parser.add_argument(
        "--include-all",
        action="store_true",
        help="Export every populated player row from each selected league sheet, ignoring --top-n.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook)
    output_path = Path(args.output)

    if not workbook_path.exists():
        raise SystemExit(f"File not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    selected_sheets = args.sheets or DEFAULT_LEAGUE_SHEETS
    missing = [sheet for sheet in selected_sheets if sheet not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Workbook is missing expected sheet(s): {missing}")

    players: list[dict[str, Any]] = []
    per_league_counts: dict[str, int] = {}

    for sheet_name in selected_sheets:
        ws = wb[sheet_name]
        rows, header_map = rows_from_sheet(ws)
        rows = sorted(rows, key=lambda r: to_int(r.get(header_map["rank"]), fallback=999999))
        if not args.include_all:
            rows = rows[: max(0, args.top_n)]
        league_players = [build_player(row, header_map, sheet_name, workbook_path.name) for row in rows]
        players.extend(league_players)
        per_league_counts[sheet_name] = len(league_players)

    output_path.write_text(json.dumps(players, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Created {output_path} with {len(players)} League Legends players")
    for league, count in per_league_counts.items():
        print(f"- {league}: {count}")


if __name__ == "__main__":
    main()
