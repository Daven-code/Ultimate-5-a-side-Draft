"""Convert the All-Time Unique Ranking tab to players_ballondor_august.json.
Usage: python convert_ballon_dor_excel_to_json.py [workbook] [output]
"""
import json, sys
from pathlib import Path
from openpyxl import load_workbook
src=Path(sys.argv[1]) if len(sys.argv)>1 else Path('Ballon_dOr_Winners.xlsx')
out=Path(sys.argv[2]) if len(sys.argv)>2 else Path('players_ballondor_august.json')
if not src.exists(): raise SystemExit(f'File not found: {src}')
wb=load_workbook(src,data_only=True,read_only=True); sheet='All-Time Unique Ranking'
if sheet not in wb.sheetnames: raise SystemExit(f'Missing sheet: {sheet}')
ws=wb[sheet]; rows=ws.iter_rows(values_only=True); headers=[str(v or '').strip() for v in next(rows)]
required=['Rank','Player','Rating_OVR','Ballon_dOr_Wins','Position','Main_Position','Nation','Position_Multiplier_DEF','Position_Multiplier_MID','Position_Multiplier_FWD']
missing=[h for h in required if h not in headers]
if missing: raise SystemExit(f'Missing columns: {missing}')
i={h:headers.index(h) for h in headers}; players=[]
for excel_row,row in enumerate(rows,2):
    name=str(row[i['Player']] or '').strip()
    if not name: continue
    def num(h):
        try:return float(row[i[h]])
        except (TypeError,ValueError):raise SystemExit(f'Invalid {h} on row {excel_row}')
    fwd=num('Position_Multiplier_FWD')
    players.append({'Player':name,'Rank':int(num('Rank')),'Game_Year':0,'Rating_OVR':int(num('Rating_OVR')),'Position':str(row[i['Position']] or '').strip(),'Main_Position':str(row[i['Main_Position']] or '').strip(),'Club':'','Nation':str(row[i['Nation']] or '').strip(),'Ballon_dOr_Wins':int(num('Ballon_dOr_Wins')),'Challenge_Mode':'BALLONDOR_AUGUST','Position_Multipliers':{'DEF':num('Position_Multiplier_DEF'),'MID':num('Position_Multiplier_MID'),'FWD':fwd,'ST':fwd}})
out.write_text(json.dumps(players,ensure_ascii=False,indent=2),encoding='utf-8')
print(f'Created {out} with {len(players)} unique Ballon d\'Or winners')
