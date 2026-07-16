"""
Convert the Ultimate 5-a-side Excel player pool into players.json for the web app.

Default workbook name:
  Ultimate5Aside_Player_Pool.xlsm
  Ultimate5Aside_Player_Pool.xlsx

Basic usage from PowerShell in this folder:
  python convert_excel_to_players_json.py

Or specify a workbook explicitly:
  python convert_excel_to_players_json.py "Ultimate5Aside_Player_Pool.xlsm"

Useful future filter examples:
  python convert_excel_to_players_json.py --league "Premier League"
  python convert_excel_to_players_json.py --club "Real Madrid" --nation "France"
  python convert_excel_to_players_json.py --year-range 2005 2017
  python convert_excel_to_players_json.py --main-position GK DEF MID FWD
  python convert_excel_to_players_json.py --challenge "WC2026"
  python convert_excel_to_players_json.py --min-rating 84
  python convert_excel_to_players_json.py --output players_premier_league.json --league "Premier League"

Notes:
- Reads the Data tab by default.
- Extra columns such as League and Challenge_Mode are included when present.
- Filters are case-insensitive.
- Multi-value filters can be supplied as repeated values, or comma-separated values.

v2 speed fix:
- Uses iter_rows(values_only=True) instead of repeated worksheet.cell(...) calls in read-only mode.
  This is much faster for .xlsm files and avoids the apparent hanging behaviour.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

DEFAULT_WORKBOOK_STEM = "Ultimate5Aside_Player_Pool"
DEFAULT_SHEET = "Data"
DEFAULT_OUTPUT = "players.json"

REQUIRED_COLUMNS = [
    "Player",
    "Game_Year",
    "Rating_OVR",
    "Position",
    "Main_Position",
    "Club",
    "Nation",
]

OPTIONAL_EXPORT_COLUMNS = [
    "League",
    "Challenge_Mode",
    "Original_Player",
    "Source_Sheet",
    "Notes",
]


def normalise_text(value: Any) -> str:
    return str(value or "").strip().casefold()


def split_filter_values(values: Iterable[str] | None) -> set[str]:
    if not values:
        return set()
    output: set[str] = set()
    for item in values:
        for part in str(item).split(","):
            clean = normalise_text(part)
            if clean:
                output.add(clean)
    return output


def find_default_workbook() -> Path:
    candidates = [
        Path(f"{DEFAULT_WORKBOOK_STEM}.xlsm"),
        Path(f"{DEFAULT_WORKBOOK_STEM}.xlsx"),
        Path(f"{DEFAULT_WORKBOOK_STEM}.xls"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate

    matches = sorted(
        list(Path.cwd().glob(f"{DEFAULT_WORKBOOK_STEM}*.xlsm"))
        + list(Path.cwd().glob(f"{DEFAULT_WORKBOOK_STEM}*.xlsx"))
    )
    if matches:
        return matches[0]

    raise SystemExit(
        f'Could not find "{DEFAULT_WORKBOOK_STEM}.xlsm" or "{DEFAULT_WORKBOOK_STEM}.xlsx" in {Path.cwd()}. '
        "Either rename the workbook or pass the file path as the first argument."
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert the Ultimate5Aside_Player_Pool workbook Data tab to players.json."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        help=f'Workbook path. If omitted, the script looks for "{DEFAULT_WORKBOOK_STEM}.xlsm/.xlsx".',
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help=f'Sheet to read. Default: "{DEFAULT_SHEET}".')
    parser.add_argument("--output", "-o", default=DEFAULT_OUTPUT, help=f'Output JSON file. Default: "{DEFAULT_OUTPUT}".')

    parser.add_argument("--league", nargs="*", help='Filter by league, e.g. --league "Premier League" "La Liga"')
    parser.add_argument("--club", nargs="*", help='Filter by club, e.g. --club "Real Madrid"')
    parser.add_argument("--nation", nargs="*", help='Filter by nation, e.g. --nation "France"')
    parser.add_argument("--position", nargs="*", help='Filter by raw position, e.g. --position ST CF CAM')
    parser.add_argument("--main-position", nargs="*", help='Filter by mapped position, e.g. --main-position GK DEF MID FWD')
    parser.add_argument("--challenge", nargs="*", help='Filter by Challenge_Mode column if present, e.g. --challenge WC2026')

    parser.add_argument("--year", nargs="*", type=int, help="Filter to specific year(s), e.g. --year 2005 2006 2026")
    parser.add_argument("--year-range", nargs=2, type=int, metavar=("START", "END"), help="Filter inclusive year range, e.g. --year-range 2005 2017")
    parser.add_argument("--min-rating", type=int, help="Minimum Rating_OVR to include.")
    parser.add_argument("--max-rating", type=int, help="Maximum Rating_OVR to include.")

    parser.add_argument("--list-filters", action="store_true", help="Print available filter values before export.")
    parser.add_argument("--pretty", action="store_true", help="Write indented JSON. Default is compact JSON.")
    return parser.parse_args()


def cell_value(row: dict[str, Any], header: str, default: Any = None) -> Any:
    value = row.get(header, default)
    return default if value is None else value


def to_int(value: Any, field_name: str, row_number: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        raise SystemExit(f"Invalid {field_name} value on worksheet row {row_number}: {value!r}")


def load_rows(workbook_path: Path, sheet_name: str) -> tuple[list[dict[str, Any]], list[str]]:
    if not workbook_path.exists():
        raise SystemExit(f"File not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f'Workbook must contain a sheet named "{sheet_name}". Found: {workbook.sheetnames}')

    worksheet = workbook[sheet_name]
    rows_iter = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise SystemExit(f'Sheet "{sheet_name}" is empty.')

    headers = [str(header).strip() if header is not None else "" for header in header_row]
    missing = [header for header in REQUIRED_COLUMNS if header not in headers]
    if missing:
        raise SystemExit(f"Missing expected columns: {missing}")

    rows: list[dict[str, Any]] = []
    for excel_row_number, values in enumerate(rows_iter, start=2):
        row = {
            headers[index]: value
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        row["__row_number"] = excel_row_number
        rows.append(row)

    workbook.close()
    return rows, headers


def row_matches_filters(row: dict[str, Any], args: argparse.Namespace) -> bool:
    if not cell_value(row, "Player"):
        return False

    text_filters = {
        "League": split_filter_values(args.league),
        "Club": split_filter_values(args.club),
        "Nation": split_filter_values(args.nation),
        "Position": split_filter_values(args.position),
        "Main_Position": split_filter_values(args.main_position),
        "Challenge_Mode": split_filter_values(args.challenge),
    }

    for column_name, allowed_values in text_filters.items():
        if allowed_values and normalise_text(row.get(column_name)) not in allowed_values:
            return False

    year = to_int(row.get("Game_Year"), "Game_Year", row["__row_number"])
    if args.year and year not in set(args.year):
        return False
    if args.year_range:
        start, end = sorted(args.year_range)
        if year < start or year > end:
            return False

    rating = to_int(row.get("Rating_OVR"), "Rating_OVR", row["__row_number"])
    if args.min_rating is not None and rating < args.min_rating:
        return False
    if args.max_rating is not None and rating > args.max_rating:
        return False

    return True


def build_player(row: dict[str, Any], headers: list[str]) -> dict[str, Any]:
    row_number = row["__row_number"]
    player = {
        "Player": str(cell_value(row, "Player", "")).strip(),
        "Game_Year": to_int(row.get("Game_Year"), "Game_Year", row_number),
        "Rating_OVR": to_int(row.get("Rating_OVR"), "Rating_OVR", row_number),
        "Position": cell_value(row, "Position", ""),
        "Main_Position": cell_value(row, "Main_Position", ""),
        "Club": cell_value(row, "Club", ""),
        "Nation": cell_value(row, "Nation", ""),
    }

    for column_name in OPTIONAL_EXPORT_COLUMNS:
        if column_name in headers:
            value = row.get(column_name)
            if value not in (None, ""):
                player[column_name] = value

    return player


def print_available_filters(rows: list[dict[str, Any]], headers: list[str]) -> None:
    print("Available filter values:")
    for column_name in ["Game_Year", "League", "Club", "Nation", "Main_Position", "Challenge_Mode"]:
        if column_name not in headers:
            continue
        values = sorted({str(row.get(column_name)).strip() for row in rows if row.get(column_name) not in (None, "")})
        preview = ", ".join(values[:40])
        suffix = "..." if len(values) > 40 else ""
        print(f"- {column_name} ({len(values)}): {preview}{suffix}")


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook) if args.workbook else find_default_workbook()

    print(f"Reading workbook: {workbook_path}")
    rows, headers = load_rows(workbook_path, args.sheet)
    print(f"Loaded {len(rows)} rows from sheet: {args.sheet}")

    if args.list_filters:
        print_available_filters(rows, headers)

    players = [build_player(row, headers) for row in rows if row_matches_filters(row, args)]
    players.sort(key=lambda p: (p["Game_Year"], -p["Rating_OVR"], str(p["Player"])))

    output_path = Path(args.output)
    json_kwargs = {"ensure_ascii": False}
    if args.pretty:
        json_kwargs["indent"] = 2
    else:
        json_kwargs["separators"] = (",", ":")

    output_path.write_text(json.dumps(players, **json_kwargs), encoding="utf-8")

    print(f"Created {output_path} with {len(players)} players")
    print("Done.")


if __name__ == "__main__":
    main()
