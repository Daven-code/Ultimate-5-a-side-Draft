"""
Convert the Excel workbook Data tab into players.json for the web app.

Usage from PowerShell in this folder:
  python -m pip install openpyxl
  python convert_excel_to_players_json.py "FIFA_Top_50_Ratings_2005_to_2026_Final.xlsx"

This creates/replaces players.json.
"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

if len(sys.argv) < 2:
    raise SystemExit('Usage: python convert_excel_to_players_json.py "your_workbook.xlsx"')

xlsx = Path(sys.argv[1])
if not xlsx.exists():
    raise SystemExit(f'File not found: {xlsx}')

wb = load_workbook(xlsx, data_only=True)
if 'Data' not in wb.sheetnames:
    raise SystemExit('Workbook must contain a sheet named Data')

ws = wb['Data']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
required = ['Player', 'Game_Year', 'Rating_OVR', 'Position', 'Main_Position', 'Club', 'Nation']
missing = [h for h in required if h not in headers]
if missing:
    raise SystemExit(f'Missing expected columns: {missing}')

idx = {h: headers.index(h) + 1 for h in required}
players = []
for r in range(2, ws.max_row + 1):
    player = ws.cell(r, idx['Player']).value
    if not player:
        continue
    players.append({
        'Player': player,
        'Game_Year': int(ws.cell(r, idx['Game_Year']).value),
        'Rating_OVR': int(ws.cell(r, idx['Rating_OVR']).value),
        'Position': ws.cell(r, idx['Position']).value,
        'Main_Position': ws.cell(r, idx['Main_Position']).value,
        'Club': ws.cell(r, idx['Club']).value,
        'Nation': ws.cell(r, idx['Nation']).value,
    })

Path('players.json').write_text(json.dumps(players, ensure_ascii=False, indent=2), encoding='utf-8')
print(f'Created players.json with {len(players)} players')
